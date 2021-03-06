#!/usr/bin/python

import sys
import os
from openpyxl import Workbook
from openpyxl import load_workbook

class StringCheck:
    
    _androidStringArr = []
    _iosStringArr = []
    _localizationPathArr = []
    _localizationDictArr = []
    #wb = Workbook()
    #ws = wb.active
    #ws['A1'] = 42
    #ws.append([1,2,3])
    #import datetime
    #ws['A2'] = datetime.datetime.now()
    #wb.save("sample.xlsx")
    
    #print(sheetTap_android['C1'].value)
    
    def makeArr(self, wb, tapname, array):
        print "make" + tapname
        sheetTap = wb[tapname]
        rowCount = 0;
        for row in sheetTap.rows:
            array.append([])
            for cell in row:
                if cell.value != None:
                    if type(cell.value) == unicode:
                        endcodeString = cell.value.encode('utf-8')
                        uniString = unicode(endcodeString,'utf-8')
                        array[rowCount].append(uniString)
                        #print str(rowCount) + str(uniString.encode('utf-8'))
                    else:
                        array[rowCount].append(cell.value)
                        #print str(rowCount) + str(uniString.encode('utf-8'))
                else:
                    array[rowCount].append("&&None&&")
            rowCount+=1
        print len(array)
    
    def loadString(self, filePath):
        print filePath
        _wb = load_workbook(filename = filePath, read_only=True)
        self._androidStringArr = []
        self.makeArr(_wb,'Android-New',self._androidStringArr)
        self._iosStringArr = []
        self.makeArr(_wb,'iOS-New',self._iosStringArr)
        print "loadString"
    
    def writeFile(self,path,line,string,sameClass,insertLast):
        f = open(path,'r')
        lines = f.readlines()
        f.close()
        if insertLast == '0' :
            #print 'insert last'
            lines.append('\n')
            lines.append(string)
        else :
            if sameClass == '0' :
                insertLine = int(line)-1
                checkString = lines[insertLine-1]
                #print 'same'
                if checkString.find('\n') < 0 :
                    lines.insert(insertLine,'\n')
                    insertLine += 1
                lines.insert(insertLine,string)
                insertLine+=1
                lines.insert(insertLine,'\n')
            else :
                insertLine = int(line)
                if insertLine > len(lines) :
                    checkString = lines[len(lines)-1]
                    if checkString != '\n' :
                        lines.append('\n')
                    lines.append('\n')
                    lines.append(string)
                else :
                    if lines[insertLine-1] != '\n':
                        lines.insert(insertLine,'\n')
                        insertLine+=1
                        lines.insert(insertLine,string)
                    else:
                        lines.insert(insertLine,string)
                    insertLine+=1
                    lines.insert(insertLine,'\n')
                    if lines[insertLine+1] != '\n':
                        insertLine+=1
                        lines.insert(insertLine,'\n')
        f = open(path,'w')
        f.writelines(lines)
        f.close()
                        
    def replaceLine(self,path,line,name,keyword,string):
        f = open(path,'r')
        lines = f.readlines()
        f.close()
        readLine = lines[int(line)-1]
        range = readLine.find('=')
        keyString = name + ':' + keyword
        if range > 0:
            key = self.getValue(readLine[:range])
            if key == keyString:
                #print string
                lines[int(line)-1] = string
                f = open(path,'w')
                f.writelines(lines)
                f.close()

    def replaceLocalString(self, line, lang, name, keyword, value, sameClass, insertLast):
        replaceString = '"' + name + ':' + keyword + '"="' + value + '";\n'
        for dictionary in self._localizationDictArr:
            index = self._localizationDictArr.index(dictionary)
            LSpath = self._localizationPathArr[index]
            langName = LSpath[LSpath.find('Resources/')+10:LSpath.rfind('.lproj')]
            if langName != lang:
                continue
            self.replaceLine(LSpath,line,name,keyword,replaceString)
            self._localizationDictArr[index] = self.getDictionaryInLSFile(LSpath)

    def writeLocalString(self, line, lang, name, keyword, value, sameClass, insertLast):
        writeString = '"' + name + ':' + keyword + '"="' + value + '";'
        #print str(line) + writeString
        for dictionary in self._localizationDictArr:
            index = self._localizationDictArr.index(dictionary)
            LSpath = self._localizationPathArr[index]
            langName = LSpath[LSpath.find('Resources/')+10:LSpath.rfind('.lproj')]
            if langName != lang:
                continue
            self.writeFile(LSpath,line,writeString,sameClass,insertLast)
            self._localizationDictArr[index] = self.getDictionaryInLSFile(LSpath)

    
    def searchPrefix(self, name, searchName, maxSearchLen):
        lowerName = name.lower()
        lowerSearchName = searchName.lower()
        if len(lowerSearchName) < 3 or len(lowerSearchName) < maxSearchLen :
            return -1
        if lowerName.find(lowerSearchName[:1]) < 0 :
            return -1
        location = lowerName.find(lowerSearchName)
        #print 'step1 ' + searchName + ' : ' + name + ': ' + str(location)
        if location == 0 :
            #print 'searchName : ' + searchName
            return len(searchName)
        else :
            return self.searchPrefix(name,searchName[:-1],maxSearchLen)
    
    def insertCheck(self, name, keyword, value):
        print 'insert : ' + name
        retArr = []
        for dictionary in self._localizationDictArr:
            searchClassName = ''
            searchMax = -1
            lineMax = -1
            sameLineMax = -1
            lastLine = -1
            index = self._localizationDictArr.index(dictionary)
            lang = self._localizationPathArr[index]
            lang = lang[lang.find('Resources/')+10:lang.rfind('.lproj')]
            sameFind = 0
            for key, val in dictionary.items():
                className = key[key.find(':')+1:key.rfind(':')]
                line = key[:key.find(':')]
                if int(lastLine) < int(line):
                    lastLine = line
                if len(className) == 0:
                    continue
                if className.lower() == name.lower():
                    sameFind = 1
                    if line > sameLineMax :
                        searchClassName = lang+'@#$'+key+'@#$'+val
                        sameLineMax = line
                    continue
                elif sameFind == 0 :
                    searchLength = self.searchPrefix(className,name,searchMax)
                    if searchLength < 0 :
                        continue
                    #print str(searchLength)
                    if searchLength == searchMax and line > lineMax :
                        lineMax = line
                        searchClassName = lang+'@#$'+key+'@#$'+val
                    if searchLength > searchMax :
                        searchClassName = lang+'@#$'+key+'@#$'+val
                        searchMax = searchLength
            if len(searchClassName) > 0 :
                retArr.append(searchClassName)
            else:
                retArr.append(lang+'@#$'+str(lastLine)+': : '+'@#$'+'lastLine')
        return retArr
    
    def findLangIndex(self,lang):
        langArr = self._iosStringArr[0]
        langIndex = -1
        if lang == 'ko':
            langIndex = langArr.index('Korean')
        elif lang == 'en':
            langIndex = langArr.index('English')
        elif lang == 'zh-Hans':
            langIndex = langArr.index('Chinese')
        elif lang == 'ja':
            langIndex = langArr.index('Japanese')
        elif lang == 'de':
            langIndex = langArr.index('German')
        elif lang == 'es':
            langIndex = langArr.index('Spanish')
        elif lang == 'fr':
            langIndex = langArr.index('French')
        elif lang == 'ru':
            langIndex = langArr.index('Russian')
        elif lang == 'it':
            langIndex = langArr.index('Italian')
        else:
            print 'no'
        #print str(langIndex)
        return langIndex

    def findTranslation(self,lang,keyword,value):
        #self._iosStringArr
        langIndex = self.findLangIndex(lang)
        retString = ''
        for array in self._iosStringArr:
            if len(array) > 1 :
                uniString = array[0]
                if uniString.encode('utf-8') == keyword :
                    if str(array[langIndex].encode('utf-8')) != '&&None&&' :
                        #print str(uniString.encode('utf-8')) + '-=======-' + keyword
                        #print str(array[langIndex].encode('utf-8')) + '-=======-' + value
                        retString = str(array[langIndex].encode('utf-8'))
                        return retString
                    else:
                        retString = str(array[langIndex].encode('utf-8'))
                        return retString
        return retString
    
    def LSStringCheck(self,langStandard):
        retArr = []
        standardKeyArr = []
        for dictionary in self._localizationDictArr:
            index = self._localizationDictArr.index(dictionary)
            lang = self._localizationPathArr[index]
            lang = lang[lang.find('Resources/')+10:lang.rfind('.lproj')]
            if lang == langStandard:
                for key, val in dictionary.items():
                    keyword = key[key.find(':')+1:]
                    standardKeyArr.append(keyword)
                break
        for dictionary in self._localizationDictArr:
            index = self._localizationDictArr.index(dictionary)
            lang = self._localizationPathArr[index]
            lang = lang[lang.find('Resources/')+10:lang.rfind('.lproj')]
            if lang != langStandard:
                print lang
                for standardKey in standardKeyArr:
                    find = False
                    for key, val in dictionary.items():
                        keyword = key[key.find(':')+1:]
                        if standardKey == keyword:
                            find = True
                    if find == False:
                        print standardKey


    def translationCheck(self,prefix):
        retArr = []
        for dictionary in self._localizationDictArr:
            index = self._localizationDictArr.index(dictionary)
            lang = self._localizationPathArr[index]
            lang = lang[lang.find('Resources/')+10:lang.rfind('.lproj')]
            #print lang
            #print len(dictionary.items())
            for key, val in dictionary.items():
                if val.find(prefix) > 0 :
                    keyword = key[key.find(':')+1:]
                    translateValue = self.findTranslation(lang,keyword,val)
                    if len(translateValue) > 0 and translateValue.find(prefix) < 0 :
                        retArr.append(lang+'@#$'+key+'@#$'+val+'@#$'+translateValue)
        return retArr
    
    def duplicationCheck(self, value):
        print 'check : ' + value
        retArr = []
        for dictionary in self._localizationDictArr:
            index = self._localizationDictArr.index(dictionary)
            lang = self._localizationPathArr[index]
            lang = lang[lang.find('Resources/')+10:lang.rfind('.lproj')]
            for key, val in dictionary.items():
                if val == value:
                    retArr.append(lang+'@#$'+key+'@#$'+val)
        #print 'len' + str(len(retArr))
        return retArr

    def searchKey(self, searchKey):
        print 'search : ' + searchKey
        retArr = []
        for dictionary in self._localizationDictArr:
            index = self._localizationDictArr.index(dictionary)
            lang = self._localizationPathArr[index]
            lang = lang[lang.find('Resources/')+10:lang.rfind('.lproj')]
            for key, val in dictionary.items():
                if val.find(searchKey) > 0 :
                    retArr.append(lang+'@#$'+key+'@#$'+val)
        #print 'len' + str(len(retArr))
        return retArr
    
    def getLastLine(self,name):
        retArr = []
        for dictionary in self._localizationDictArr:
            searchClassName = ''
            lastLine = -1
            sameLineMax = -1
            index = self._localizationDictArr.index(dictionary)
            lang = self._localizationPathArr[index]
            lang = lang[lang.find('Resources/')+10:lang.rfind('.lproj')]
            sameFind = 0
            for key, val in dictionary.items():
                className = key[key.find(':')+1:key.rfind(':')]
                line = key[:key.find(':')]
                if int(lastLine) < int(line):
                    lastLine = line
                if len(className) == 0:
                    continue
                if className.lower() == name.lower():
                    sameFind = 1
                    if line > sameLineMax :
                        searchClassName = lang+'@#$'+key+'@#$'+val
                        sameLineMax = line
                    continue
            if len(searchClassName) > 0 :
                retArr.append(searchClassName)
            else:
                retArr.append(lang+'@#$'+str(lastLine)+': : '+'@#$'+'lastLine')
        return retArr
    
    def is_asleep(self):
        return True
    
    def insertString(self, name, keyword, value):
        print name + keyword + value
        return self.insertCheck(name, keyword, value)
    
    
    def getValue(self,line):
        start = line.find("\"")
        end = line.rfind("\"")
        return line[start+1:end]
    
    def getDictionaryInLSFile(self, filePath):
        dictionary = {}
        f = open(filePath, 'r')
        i = 0
        while 1:
            i+=1
            line = f.readline()
            range = line.find('=')
            key = str(i)
            if range > 0:
                keyword = self.getValue(line[:range])
                value = self.getValue(line[range+1:])
                dictionary[key+':'+keyword] = value
            else :
                dictionary[key+':empty'] = 'empty'
            if not line: break
        f.close()
        return dictionary

    def loadLocalString(self, filePath):
        self._localizationPathArr = []
        self._localizationDictArr = []
        for filename in os.listdir(filePath):
            if filename.find('lproj') > 0 :
                if filename.find('zh-Hant') > 0 :
                    continue
                for lsPath in os.listdir(filePath+'/'+filename):
                    if lsPath.find('strings') > 0 :
                        self._localizationPathArr.append(filePath+'/'+filename+'/'+lsPath)
        for path in self._localizationPathArr:
            if path.find('zh-Hant') > 0 :
                continue
            print path
            dictionary = self.getDictionaryInLSFile(path)
            self._localizationDictArr.append(dictionary)

if __name__ == "__main__":
    _stringCheck = StringCheck()
    #_stringCheck.loadString()


